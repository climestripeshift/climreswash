"""
Turn the manually-compiled CSR company list into a per-district picture:
how many CSR companies are active in each of Rajasthan's 41 current
districts (with a thematic-area breakdown), versus how many schools in
that district actually have a documented infrastructure need -- so a
district with high need and zero CSR presence reads as a real gap, not
just a number.

Source: ~/UNICEF RAJASTHAN/2026/wins/CSR/CSR data Analysis -Final.xlsx
(a single sheet, one row per company: name/contact, primary district,
a free-text "CSR in other districts of Rajasthan" cell, 5 thematic-area
Yes/blank flags, budget, and an annual-report link). An identical copy
also sits in the Workshop/ subfolder -- not read, just a duplicate.

District resolution is inherently messy: the primary-district column has
typos and formatting variants (Beawer/Siker/Jhunjunu, "Jaipur(urban)" vs
"Jaipur (Urban)", "Sriganganagar" one word, etc.), and the free-text
"other districts" column mixes real district names with sub-district town
names (Neemrana, Kawai, Zawar, Nimbol...), a couple of out-of-state towns
that got mentioned alongside Rajasthan ones (Lavasa, Halol), and vague
"whole state" language ("All Rajsthan", "23 District of Rajasthan").

DISTRICT_ALIAS below is a hand-built lookup covering every distinct
raw value actually seen in this file (verified by re-running this
script and confirming the "unresolved tokens" report at the end comes
back empty) -- not a general-purpose parser. A few entries are
best-effort geographic calls flagged with a comment; if wrong, fix the
alias and rerun -- nothing downstream needs to change.

"Blanket"/statewide companies (comes from language like "All Rajsthan",
a bare "Rajasthan/Rajsthan", or a primary district that IS "Rajasthan")
are counted toward every one of the 41 districts, but kept in a
separate count from district-specific ones so a district's number isn't
inflated by companies with no actual local presence -- see
csr_specific_count vs csr_statewide_count in the output.

"23 District of Rajasthan" / "Other districts of Rajasthan" language is
NOT treated as a full blanket (we don't know which 23) -- companies
using that language are counted in csr_broad_unspecified_count instead,
kept separate from both the specific and statewide buckets.

Run: python scripts/build_csr_district_map.py
"""
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

from join_shvr_infrastructure_needs import canonical_district

ROOT = Path(__file__).resolve().parent.parent
SRC = Path.home() / "UNICEF RAJASTHAN/2026/wins/CSR/CSR data Analysis -Final.xlsx"
SCHOOLS = ROOT / "client/public/data/shvr_schools_infra_rajasthan.json"
DISTRICTS_GEO = ROOT / "client/public/data/rajasthan_districts_current.geojson"

OUT_COMPANIES = ROOT / "client/public/data/csr_companies_rajasthan.json"
OUT_DISTRICT = ROOT / "client/public/data/csr_district_summary_rajasthan.json"

THEMES = [
    ("formal_education", "Thematic area\nFormal Education"),
    ("wash", "Thematic area\nWASH"),
    ("school_hardware", "Thematic area\nSchool hardware"),
    ("other_school_initiative", "Thematic area\nOther school initiative (ICT, sports etc.)"),
    ("anganwadi", "Thematic area\nAnganwadi"),
]

# raw (uppercased, whitespace-collapsed) token -> current-district name, or one of the two
# sentinels below. Covers both the primary-district column and every token the "other
# districts" free text splits into (see split_other_districts).
BLANKET = "__BLANKET__"          # explicit "whole state" language
OUT_OF_STATE = "__OUT_OF_STATE__"  # a real place, just not in Rajasthan

DISTRICT_ALIAS: dict[str, str] = {
    # typos / formatting seen in the primary-district column (matched after clean_key()
    # strips ALL punctuation, not just leading/trailing, so "Jaipur (Urban)" and
    # "Jaipur(urban)" both normalize to the same "JAIPUR URBAN" key)
    "BEAWER": "Beawar", "SIKER": "Sikar", "JHUNJUNU": "Jhunjhunun",
    "KHARTHAL TIJARA": "Khairthal-Tijara", "JAIPUR RURAL": "Jaipur", "JAIPUR URBAN": "Jaipur",
    "CHITTORGARH RAJASTHAN": "Chittaurgarh", "SRIGANGANAGAR": "Ganganagar",
    "ALVAR": "Alwar", "BARN": "Baran", "BEAVER": "Beawar", "KAROLI": "Karauli",
    "SALUMBER": "Salumbar", "SRI GANGANAGAR": "Ganganagar",
    # Bhiwadi is an industrial town, not one of the 41 current districts -- mapping it to
    # its historical parent district Alwar (not the newer Khairthal-Tijara split, which
    # covers Alwar's NE tehsils bordering Haryana). Best-effort; flagged in the summary
    # meta for manual review since several respondents list "Alwar" and "Bhiwadi" as if
    # they were two different places.
    "BHIWADI": "Alwar", "KHIJURIWAS": "Alwar",
    # descriptive context that split off its own token alongside an already-resolved one
    # (e.g. "Bhiwadi (Delhi-NCR)" splits into "Bhiwadi" + "Delhi-NCR") -- not place claims
    # of their own, so treated as junk rather than "unresolved"
    "DELHI NCR": None, "BHIWADI ALWAR ROAD": None, "BEAWAR JAITARAN SIDE": None,
    # sub-district towns named in the free-text "other districts" column, resolved to
    # their district (only towns actually seen in the source data -- not a gazetteer)
    "ABU ROAD": "Sirohi", "KAWAI": "Baran", "NEEMRANA": "Alwar", "ZAWAR": "Udaipur",
    "RAWATBHATA REGION": "Chittaurgarh", "RAWATBHATA": "Chittaurgarh",
    "GOTAN NAGAUR": "Nagaur", "GOTAN": "Nagaur",
    "MANGROL CHITTORGARH": "Chittaurgarh", "MANGROL": "Chittaurgarh",
    "NIMBAHERA CHITTORGARH": "Chittaurgarh", "NIMBAHERA": "Chittaurgarh",
    "NIMBOL": "Pali", "BHAWANIMANDI": "Jhalawar", "JHALAWAR DISTRICT": "Jhalawar",
    # explicit "whole state" language
    "RAJASTHAN": BLANKET, "RAJSTHAN": BLANKET, "ALL RAJSTHAN": BLANKET,
    "ALL RAJASTHAN": BLANKET,
    # real places, just not in Rajasthan (mentioned alongside RJ ones in a couple of rows)
    "GUJARAT": OUT_OF_STATE, "ANDHRA PRADESH": OUT_OF_STATE,
    "LAVASA": OUT_OF_STATE, "HALOL": OUT_OF_STATE,
}
# free-standing junk tokens from splitting messy punctuation -- dropped silently, not
# logged as "unresolved" (they're not place names at all)
JUNK_TOKENS = {"N", ""}
# "23 District of Rajasthan" / "Other districts of Rajasthan" -- broad but not a full
# blanket claim (we don't know which districts), and not resolvable to one district either
BROAD_UNSPECIFIED_PATTERNS = [
    re.compile(r"^\d+\s*DISTRICT", re.I),
    re.compile(r"OTHER DISTRICT", re.I),
]


def clean_key(v: str) -> str:
    """Upper-case with ALL punctuation (not just leading/trailing) stripped, so
    "Jaipur (Urban)" and "Jaipur(urban)" both normalize to "JAIPUR URBAN"."""
    v = re.sub(r"[(),./-]", " ", str(v or "").upper())
    return re.sub(r"\s+", " ", v).strip()


def resolve_token(raw: str, current_districts: set[str]) -> tuple[str | None, str]:
    """Returns (resolved_district_or_None, kind) where kind is one of
    'district' / 'blanket' / 'broad' / 'out_of_state' / 'unresolved' / 'junk'."""
    key = clean_key(raw)
    if key in JUNK_TOKENS:
        return None, "junk"
    if key in DISTRICT_ALIAS:
        v = DISTRICT_ALIAS[key]
        if v is None:
            return None, "junk"
        if v == BLANKET:
            return None, "blanket"
        if v == OUT_OF_STATE:
            return None, "out_of_state"
        return v, "district"
    for pat in BROAD_UNSPECIFIED_PATTERNS:
        if pat.search(key):
            return None, "broad"
    canon = canonical_district(raw.strip())
    if canon in current_districts:
        return canon, "district"
    return None, "unresolved"


def split_other_districts(raw: str) -> list[str]:
    # split on comma, newline, "and", "&", and parens (a couple of rows use
    # "Rajasthan (Jaipur, Udaipur)" style parenthetical lists)
    parts = re.split(r"[,\n()]|(?:\s+and\s+)|(?:\s*&\s*)", str(raw))
    return [p.strip() for p in parts if p.strip()]


def yes(v) -> bool:
    return str(v or "").strip().upper() == "YES"


def main():
    current_districts = {f["properties"]["NAME"] for f in json.loads(DISTRICTS_GEO.read_text())["features"]}

    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Final "]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]
    col = {h: i for i, h in enumerate(headers)}

    companies = []
    unresolved_tokens: dict[str, int] = defaultdict(int)
    broad_unspecified_companies = 0

    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 14)]
        if not vals[1]:  # no company name -> blank padding row
            continue
        name = str(vals[1]).strip()
        primary_raw = vals[col["District in Rajasthan "]]
        other_raw = vals[col["CSR in Other districts of Rajasthan "]]

        districts: set[str] = set()
        is_blanket = False
        is_broad_unspecified = False

        if primary_raw:
            d, kind = resolve_token(str(primary_raw), current_districts)
            if kind == "district":
                districts.add(d)
            elif kind == "blanket":
                is_blanket = True
            elif kind == "broad":
                is_broad_unspecified = True
            elif kind == "unresolved":
                unresolved_tokens[str(primary_raw).strip()] += 1

        if other_raw:
            for token in split_other_districts(other_raw):
                d, kind = resolve_token(token, current_districts)
                if kind == "district":
                    districts.add(d)
                elif kind == "blanket":
                    is_blanket = True
                elif kind == "broad":
                    is_broad_unspecified = True
                elif kind == "unresolved":
                    unresolved_tokens[token] += 1
                # 'junk' and 'out_of_state' tokens are silently dropped

        if is_broad_unspecified and not districts and not is_blanket:
            broad_unspecified_companies += 1

        themes = {key: yes(vals[col[header]]) for key, header in THEMES}
        budget = vals[col["Budget spent during last year or earlier on the thematic area"]]
        report_link = vals[col["Annual report link"]]

        companies.append({
            "name": name,
            "contact_person": vals[col["Concerned person's Name"]],
            "contact_info": vals[col["Mobile/Email"]],
            "primary_district_raw": primary_raw,
            "other_districts_raw": other_raw,
            "districts": sorted(districts),
            "is_statewide": is_blanket,
            "is_broad_unspecified": is_broad_unspecified and not districts and not is_blanket,
            "themes": themes,
            "budget_raw": budget,
            "annual_report_link": report_link,
        })

    print(f"Parsed {len(companies)} companies")
    statewide = sum(1 for c in companies if c["is_statewide"])
    broad = sum(1 for c in companies if c["is_broad_unspecified"])
    no_district = sum(1 for c in companies if not c["districts"] and not c["is_statewide"] and not c["is_broad_unspecified"])
    print(f"  {statewide} statewide/blanket, {broad} broad-unspecified (\"23 districts\" etc.), "
          f"{no_district} with no resolved district at all")

    if unresolved_tokens:
        print(f"\nUnresolved tokens ({len(unresolved_tokens)} distinct) -- add these to DISTRICT_ALIAS "
              f"if they're real places, then rerun:")
        for t, c in sorted(unresolved_tokens.items(), key=lambda kv: -kv[1]):
            print(f"  {t!r}: seen {c}x")
    else:
        print("\nNo unresolved tokens -- every raw district/other-district value resolved cleanly.")

    # ── per-district CSR summary ──────────────────────────────────────────────
    by_district_companies: dict[str, list[dict]] = defaultdict(list)
    for c in companies:
        for d in c["districts"]:
            by_district_companies[d].append(c)

    statewide_companies = [c for c in companies if c["is_statewide"]]

    # ── cross-reference against the school infra-need registry ─────────────────
    schools = json.loads(SCHOOLS.read_text())
    need_by_district: dict[str, dict] = defaultdict(lambda: {
        "total_school_count": 0, "schools_needing_help_count": 0,
        "toilet_required_count": 0, "classroom_repair_needed_count": 0,
        "building_dilapidated_count": 0, "new_classroom_requirement_count": 0,
    })
    for s in schools:
        # matches join_shvr_infrastructure_needs.py's own district grouping exactly:
        # district_raw is the finer-grained UDISE text (already reflects the 2023 district
        # split, e.g. distinguishes Beawar from Ajmer); the plain "district" field is
        # coarser/legacy and undercounts newer districts if used directly.
        raw = s.get("district_raw") or s.get("district")
        if not raw:
            continue
        d = canonical_district(raw)
        if d not in current_districts:
            continue
        n = need_by_district[d]
        n["total_school_count"] += 1
        # girls_toilet_required is a truthy toilet-unit COUNT (int), not a bool -- None/0
        # means no requirement recorded, same convention join_shvr_infrastructure_needs.py uses
        toilet = bool(s.get("girls_toilet_required"))
        repair = s.get("classroom_repair_needed") is True
        dilapidated = s.get("building_dilapidated") is True
        new_classroom = s.get("new_classroom_requirement") is True
        if toilet:
            n["toilet_required_count"] += 1
        if repair:
            n["classroom_repair_needed_count"] += 1
        if dilapidated:
            n["building_dilapidated_count"] += 1
        if new_classroom:
            n["new_classroom_requirement_count"] += 1
        if toilet or repair or dilapidated or new_classroom:
            n["schools_needing_help_count"] += 1

    all_districts = sorted(current_districts)
    by_district = {}
    for d in all_districts:
        specific = by_district_companies.get(d, [])
        need = need_by_district.get(d, {
            "total_school_count": 0, "schools_needing_help_count": 0,
            "toilet_required_count": 0, "classroom_repair_needed_count": 0,
            "building_dilapidated_count": 0, "new_classroom_requirement_count": 0,
        })
        theme_counts = {key: sum(1 for c in specific if c["themes"][key]) for key, _ in THEMES}
        theme_counts_statewide = {key: sum(1 for c in statewide_companies if c["themes"][key]) for key, _ in THEMES}
        by_district[d] = {
            "csr_specific_count": len(specific),
            "csr_specific_companies": sorted({c["name"] for c in specific}),
            "csr_statewide_count": len(statewide_companies),
            "csr_total_available_count": len(specific) + len(statewide_companies),
            "csr_specific_by_theme": theme_counts,
            "csr_statewide_by_theme": theme_counts_statewide,
            **need,
            "needs_help_no_specific_csr": need["schools_needing_help_count"] > 0 and len(specific) == 0,
        }

    OUT_COMPANIES.write_text(json.dumps(companies, indent=2, default=str))
    OUT_DISTRICT.write_text(json.dumps({
        "meta": {
            "source": str(SRC),
            "note": "District resolution of the free-text \"CSR in other districts\" column is "
                    "best-effort (see DISTRICT_ALIAS in build_csr_district_map.py) -- Bhiwadi in "
                    "particular is mapped to Alwar as a judgment call, not a confirmed admin fact. "
                    "csr_statewide_count is the same number for every district (companies whose "
                    "coverage language reads as \"all of Rajasthan\"); csr_broad_unspecified "
                    "companies (\"23 districts\", unspecified) aren't counted toward any single "
                    "district since which ones isn't stated.",
            "total_companies": len(companies),
            "statewide_companies": statewide,
            "broad_unspecified_companies": broad,
            "unresolved_tokens": dict(unresolved_tokens),
            "generated_from_rows": len(companies),
        },
        "by_district": by_district,
    }, indent=2))

    print(f"\nSaved {OUT_COMPANIES}")
    print(f"Saved {OUT_DISTRICT}")

    print("\nDistricts with need but zero specific CSR presence (statewide-only or nothing):")
    gap = [(d, v) for d, v in by_district.items() if v["needs_help_no_specific_csr"]]
    gap.sort(key=lambda kv: -kv[1]["schools_needing_help_count"])
    for d, v in gap[:15]:
        print(f"  {d:20s} {v['schools_needing_help_count']:5d} schools need help, "
              f"0 specific CSR (+{v['csr_statewide_count']} statewide)")


if __name__ == "__main__":
    main()

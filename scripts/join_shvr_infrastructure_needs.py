"""
Join UNICEF Rajasthan CSR infrastructure-needs lists (new classroom
requirement, classroom repair, dilapidated buildings, toilet requirement)
onto a UNIFIED school registry — the union of SHVR-rated schools and
every school appearing in any of the 4 CSR files, not just the overlap.

Source files: ~/UNICEF RAJASTHAN/2026/wins/CSR/fwd_/*.xlsx (not committed —
personal UNICEF working files, outside the repo)

Why a union, not a join onto SHVR alone: the 3 UDISE-keyed CSR files
together cover 55,387 distinct schools, but only 26,712 of those overlap
with SHVR's 39,602. That leaves 28,675 real schools with a documented
infrastructure need (repair/new classroom/dilapidated) that would be
invisible if this script only ever looked schools up by SHVR UDISE code.
Symmetrically, ~12,890 SHVR-rated schools never appear in any CSR list —
which is a real "no need" signal, not the CSR files failing to cover them.

Three of the four CSR files join reliably by UDISE code. Toilet
requirement has a broken export where every row shares the same UDISE
code, so it's matched by school name within district instead (against
the full unified registry, not just SHVR) — lower confidence, flagged
per-school via toilet_match_method.

Data-integrity guard: even in the "reliable" files, a UDISE code that
maps to genuinely different school names (a smaller instance of the
same export bug) is dropped rather than trusted — a rare mislabeled
row is worse than a missing one.

New-to-the-registry schools (CSR-only, no SHVR entry) get no rating,
status "not_in_shvr", and no village-matched location (CSR files carry
no address text to match against) — they render at their district's
centroid, same as any other unlocated school.

Run: python scripts/join_shvr_infrastructure_needs.py
"""
import difflib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC  = Path.home() / "UNICEF RAJASTHAN/2026/wins/CSR/fwd_"
# Reads its own prior output as input — safe/idempotent, since this script only ever
# overwrites the infra_* keys (and appends new CSR-only school records) it itself
# manages, never touching the location/rating fields from the village-matching step.
SCHOOLS = ROOT / "client/public/data/shvr_schools_infra_rajasthan.json"

OUT_SCHOOLS = ROOT / "client/public/data/shvr_schools_infra_rajasthan.json"
OUT_SUMMARY = ROOT / "client/public/data/shvr_infra_by_rating_rajasthan.json"
OUT_DISTRICT = ROOT / "client/public/data/shvr_district_absolute_infra_rajasthan.json"

NAME_MATCH_THRESHOLD = 0.82

# new-district (as CSR/SHVR spell it) -> old/undivided parent (matches the hex grid,
# needed so CSR-only schools without a village match still get a district-centroid hex).
NEW_TO_OLD_DISTRICT = {
    "Balotara": "Barmer", "Beawer": "Ajmer", "Beawar": "Ajmer", "Deeg": "Bharatpur",
    "Didwana-Kuchaman": "Nagaur", "Khairthal-Tijara": "Alwar", "Kotputli-Behror": "Jaipur",
    "Phalodi": "Jodhpur", "Salumbar": "Udaipur",
}


def norm_code(v) -> str:
    return str(v).strip().zfill(11)


def norm_name(v: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(v or "").upper())


def yes(v) -> bool:
    return str(v or "").strip().upper() == "YES"


# CSR district-column spelling variants -> canonical name (matches rajasthan_districts_current.geojson)
DISTRICT_ALIAS = {
    "BALOTRA": "Balotara", "CHITTORGARH": "Chittaurgarh", "JALORE": "Jalor",
    "DHOLPUR": "Dhaulpur", "SRI GANGANAGAR": "Ganganagar", "JHUNJHUNU": "Jhunjhunun",
    "KOTPUTLI- BEHROR": "Kotputli-Behror", "KOTPUTLY/BEHROD": "Kotputli-Behror",
    "DIDWANA - KUCHAMAN": "Didwana-Kuchaman", "SALUMBER": "Salumbar",
}


def canonical_district(raw: str) -> str:
    up = (raw or "").replace("(RAJ.)", "").strip().upper()
    return DISTRICT_ALIAS.get(up, up.title())


def clean_udise_map(rows: list[tuple[str, str, str, dict]]) -> dict[str, dict]:
    """rows: (udise_code, school_name, district_raw, data) — drops codes whose rows disagree
    on school name. Returns clean UDISE -> {**merged_data, "_name": ..., "_district": ...}."""
    by_code: dict[str, list[tuple[str, str, dict]]] = defaultdict(list)
    for code, name, district, data in rows:
        by_code[code].append((name, district, data))

    clean: dict[str, dict] = {}
    dropped = 0
    for code, entries in by_code.items():
        names = {norm_name(n) for n, _, _ in entries}
        if len(names) > 1:
            dropped += 1
            continue
        # merge all rows for this (legitimately single) school: bools OR together,
        # numbers sum (e.g. two repair line-items' classroom counts really do add up)
        merged: dict = {}
        for _, _, data in entries:
            for k, v in data.items():
                if isinstance(v, bool):
                    merged[k] = merged.get(k, False) or v
                elif isinstance(v, (int, float)) and k in merged:
                    merged[k] += v
                else:
                    merged[k] = v
        merged["_name"] = entries[0][0]
        merged["_district"] = canonical_district(entries[0][1])
        clean[code] = merged
    print(f"    {len(by_code)} distinct codes -> {len(clean)} clean, {dropped} dropped (name mismatch)")
    return clean


def load_acr_new() -> dict[str, dict]:
    print("Loading ACR_New_Requirement_List.xlsx (new classroom / dilapidated building)...")
    wb = openpyxl.load_workbook(SRC / "ACR_New_Requirement_List.xlsx", data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[5] is None:
            continue
        rows.append((norm_code(row[5]), row[4], row[1], {
            "building_fully_dilapidated": yes(row[8]),
            "dilapidated_declared_official": yes(row[9]),
            "dilapidated_classroom_count": row[10] if isinstance(row[10], (int, float)) else 0,
            "new_classroom_requirement": True,
        }))
    return clean_udise_map(rows)


def load_acr_repair() -> dict[str, dict]:
    print("Loading ACR_Raiparing List.xlsx (classroom repair)...")
    wb = openpyxl.load_workbook(SRC / "ACR_Raiparing List.xlsx", data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[5] is None:
            continue
        rows.append((norm_code(row[5]), row[4], row[1], {
            "classrooms_needing_repair": row[8] if isinstance(row[8], (int, float)) else 0,
            "repair_amount_needed_lakh": row[9] if isinstance(row[9], (int, float)) else 0,
            "classroom_repair_needed": True,
        }))
    return clean_udise_map(rows)


def load_dilapidated() -> dict[str, dict]:
    print("Loading Dilapited_Building.xlsx...")
    wb = openpyxl.load_workbook(SRC / "Dilapited_Building.xlsx", data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[4] is None:
            continue
        rows.append((norm_code(row[4]), row[3], row[1], {"building_dilapidated_listed": yes(row[6])}))
    return clean_udise_map(rows)


def load_toilet_rows() -> list[dict]:
    """UDISE codes are broken (all identical) — matched by normalized name within district."""
    print("Loading Toilet_Requirement_List.xlsx (UDISE broken, matching by name)...")
    wb = openpyxl.load_workbook(SRC / "Toilet_Requirement_List.xlsx", data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[4] is None:
            continue
        rows.append({"district": canonical_district(row[1]), "school_name": row[4], "girls_toilet_required": row[6] or 1})
    print(f"    {len(rows)} toilet-requirement rows")
    return rows


def main():
    if not SRC.exists():
        print(f"Source folder not found: {SRC}")
        return

    acr_new = load_acr_new()
    acr_repair = load_acr_repair()
    dilapidated = load_dilapidated()
    toilet_rows = load_toilet_rows()

    print(f"\nLoading {SCHOOLS}...")
    schools = json.loads(SCHOOLS.read_text())
    for s in schools:
        s.setdefault("in_shvr", True)  # preserves False on CSR-only schools from a prior run
    existing_codes = {norm_code(s["udise_code"]) for s in schools if s.get("udise_code")}
    print(f"  {len(schools)} schools already on file ({existing_codes.__len__()} unique UDISE codes, "
          f"{sum(1 for s in schools if s['in_shvr'])} from SHVR)")

    # ── Build the UNION registry: add every CSR-file school not already on file ──────
    all_csr_codes = set(acr_new) | set(acr_repair) | set(dilapidated)
    new_codes = all_csr_codes - existing_codes
    print(f"\nCSR files cover {len(all_csr_codes)} unique schools; {len(new_codes)} are new — adding them")
    for code in sorted(new_codes):
        source = acr_new.get(code) or acr_repair.get(code) or dilapidated.get(code)
        district = source["_district"]
        schools.append({
            "name": source["_name"], "udise_code": code, "address": None,
            "school_type": None, "nep_category": None,
            "district_raw": district, "district": NEW_TO_OLD_DISTRICT.get(district, district),
            "status": "not_in_shvr", "rating": None, "percentage": None,
            "lat": None, "lon": None, "h3_id": None,
            "matched_village": None, "location_precision": "district_centroid",
            "in_shvr": False,
        })
    print(f"Unified registry: {len(schools)} schools ({len(existing_codes)} on file already + {len(new_codes)} newly added)")

    # ── UDISE-based joins, over the full unified registry ───────────────────────────
    matched_new = matched_repair = matched_dilap = 0
    for s in schools:
        code = norm_code(s["udise_code"]) if s.get("udise_code") else None
        s["new_classroom_requirement"] = False
        s["classroom_repair_needed"] = False
        s["building_dilapidated"] = False
        s["dilapidated_classroom_count"] = 0
        s["classrooms_needing_repair"] = 0
        s["girls_toilet_required"] = None  # filled below, name-matched
        s["toilet_match_method"] = None

        if code and code in acr_new:
            matched_new += 1
            d = acr_new[code]
            s["new_classroom_requirement"] = True
            s["dilapidated_classroom_count"] = d["dilapidated_classroom_count"]
            if d["building_fully_dilapidated"]:
                s["building_dilapidated"] = True
        if code and code in acr_repair:
            matched_repair += 1
            d = acr_repair[code]
            s["classroom_repair_needed"] = True
            s["classrooms_needing_repair"] = d["classrooms_needing_repair"]
        if code and code in dilapidated:
            matched_dilap += 1
            s["building_dilapidated"] = True

    print(f"\nUDISE-joined: new_classroom={matched_new} repair={matched_repair} dilapidated={matched_dilap}")

    # ── Name-based join for toilet requirement, district-scoped, full registry ──────
    by_current_district: dict[str, list[dict]] = defaultdict(list)
    for s in schools:
        d = canonical_district(s["district_raw"]) if s.get("district_raw") else canonical_district(s["district"])
        by_current_district[d].append(s)

    toilet_matched = 0
    for row in toilet_rows:
        candidates = by_current_district.get(row["district"], [])
        if not candidates:
            continue
        target_norm = norm_name(row["school_name"])
        best, best_score = None, 0.0
        for s in candidates:
            score = difflib.SequenceMatcher(None, target_norm, norm_name(s["name"])).ratio()
            if score > best_score:
                best, best_score = s, score
        if best is not None and best_score >= NAME_MATCH_THRESHOLD and best["girls_toilet_required"] is None:
            best["girls_toilet_required"] = row["girls_toilet_required"]
            best["toilet_match_method"] = "name_match"
            toilet_matched += 1

    print(f"Toilet requirement matched by name: {toilet_matched}/{len(toilet_rows)} "
          f"({100*toilet_matched/len(toilet_rows):.1f}%, threshold={NAME_MATCH_THRESHOLD})")

    # ── Summary by star rating (SHVR participants only — rating doesn't exist otherwise) ──
    by_rating: dict[str, dict] = {}
    shvr_schools = [s for s in schools if s["in_shvr"]]
    for r in [1, 2, 3, 4, 5, None]:
        key = str(r) if r is not None else "unrated"
        group = [s for s in shvr_schools if s["rating"] == r]
        n = len(group)
        by_rating[key] = {
            "school_count": n,
            "toilet_required_count": sum(1 for s in group if s["girls_toilet_required"]),
            "classroom_repair_needed_count": sum(1 for s in group if s["classroom_repair_needed"]),
            "new_classroom_requirement_count": sum(1 for s in group if s["new_classroom_requirement"]),
            "building_dilapidated_count": sum(1 for s in group if s["building_dilapidated"]),
        }
        if n:
            by_rating[key]["toilet_required_pct"] = round(100 * by_rating[key]["toilet_required_count"] / n, 1)
            by_rating[key]["classroom_repair_pct"] = round(100 * by_rating[key]["classroom_repair_needed_count"] / n, 1)
            by_rating[key]["new_classroom_pct"] = round(100 * by_rating[key]["new_classroom_requirement_count"] / n, 1)
            by_rating[key]["dilapidated_pct"] = round(100 * by_rating[key]["building_dilapidated_count"] / n, 1)

    print("\nInfrastructure needs by SHVR rating:")
    for k, v in by_rating.items():
        print(f"  {k:8s} n={v['school_count']:6d}  toilet={v.get('toilet_required_pct','-'):>5}%  "
              f"repair={v.get('classroom_repair_pct','-'):>5}%  new_room={v.get('new_classroom_pct','-'):>5}%  "
              f"dilapidated={v.get('dilapidated_pct','-'):>5}%")

    # ── Summary by CURRENT district, full unified registry — real percentages now,
    # since the denominator is every known school (SHVR + CSR-only), not just raters ──
    by_district: dict[str, dict] = {}
    for d, group in sorted(by_current_district.items()):
        n = len(group)
        in_shvr = [s for s in group if s["in_shvr"]]
        rated = [s["rating"] for s in in_shvr if s["rating"] is not None]
        toilet_c = sum(1 for s in group if s["girls_toilet_required"])
        repair_c = sum(1 for s in group if s["classroom_repair_needed"])
        new_room_c = sum(1 for s in group if s["new_classroom_requirement"])
        dilap_c = sum(1 for s in group if s["building_dilapidated"])
        by_district[d] = {
            "total_school_count": n,
            "in_shvr_count": len(in_shvr),
            "has_shvr_rating": len(in_shvr) > 0,
            "avg_rating": round(sum(rated) / len(rated), 2) if rated else None,
            "toilet_required_count": toilet_c,
            "classroom_repair_needed_count": repair_c,
            "new_classroom_requirement_count": new_room_c,
            "building_dilapidated_count": dilap_c,
            "toilet_required_pct": round(100 * toilet_c / n, 1) if n else None,
            "classroom_repair_pct": round(100 * repair_c / n, 1) if n else None,
            "new_classroom_pct": round(100 * new_room_c / n, 1) if n else None,
            "dilapidated_pct": round(100 * dilap_c / n, 1) if n else None,
        }

    print(f"\nInfrastructure needs by district ({len(by_district)} districts, full registry):")
    for d, v in sorted(by_district.items(), key=lambda kv: -kv[1]["classroom_repair_needed_count"]):
        shvr_tag = f"{v['in_shvr_count']}/{v['total_school_count']} in SHVR" if v["has_shvr_rating"] else "no SHVR"
        print(f"  {d:18s} n={v['total_school_count']:6d} ({shvr_tag:18s})  repair={v['classroom_repair_needed_count']:5d} ({v['classroom_repair_pct']:>5}%)  "
              f"new_room={v['new_classroom_requirement_count']:5d}  dilapidated={v['building_dilapidated_count']:4d}  toilet={v['toilet_required_count']:4d}")

    OUT_SCHOOLS.write_text(json.dumps(schools, separators=(",", ":")))
    OUT_SUMMARY.write_text(json.dumps({
        "meta": {
            "sources": ["ACR_New_Requirement_List.xlsx", "ACR_Raiparing List.xlsx",
                        "Dilapited_Building.xlsx", "Toilet_Requirement_List.xlsx"],
            "note": "SHVR-rated schools only (rating breakdown doesn't apply to schools that "
                    "never participated in SHVR). First 3 CSR files joined by UDISE code, dropping "
                    "any code whose rows disagree on school name. Toilet requirement matched by "
                    f"normalized school name within district, threshold {NAME_MATCH_THRESHOLD} — "
                    "lower confidence than the UDISE-joined fields.",
        },
        "by_rating": by_rating,
    }, indent=2))

    OUT_DISTRICT.write_text(json.dumps({
        "meta": {
            "note": "Full unified registry (SHVR-rated schools UNION every school appearing in "
                    "any of the 4 CSR files) by CURRENT (post-2023-split) district. total_school_count "
                    "is the real denominator behind each _pct field — includes schools SHVR never "
                    "rated, so these percentages and counts are meaningful for all 41 current "
                    "districts, not just the 26 SHVR covered. has_shvr_rating / in_shvr_count show "
                    "how much of that total actually has a star rating.",
        },
        "by_district": by_district,
    }, indent=2))

    import os
    print(f"\nSaved {OUT_SCHOOLS} ({os.path.getsize(OUT_SCHOOLS)//1024}KB)")
    print(f"Saved {OUT_DISTRICT}")
    print(f"Saved {OUT_SUMMARY}")


if __name__ == "__main__":
    main()
